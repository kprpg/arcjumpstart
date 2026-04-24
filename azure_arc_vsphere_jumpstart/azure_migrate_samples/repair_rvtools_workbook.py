from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DEFAULT_TEMPLATE = BASE_DIR / "azure_migrate_vmware_rvtools_strict_sample.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)

SHEET_SCHEMA = {
    "vInfo": [
        "VM",
        "VM UUID",
        "Powerstate",
        "CPUs",
        "Memory",
        "Provisioned MiB",
        "In use MiB",
        "OS according to the configuration file",
    ],
    "vHost": [
        "Host",
        "Cluster",
        "Datacenter",
        "Config status",
        "in Maintenance Mode",
        "in Quarantine Mode",
        "CPU Model",
        "Speed",
        "#CPU",
        "Cores per CPU",
        "# Cores",
        "CPU usage %",
        "# Memory",
        "Memory usage %",
        "VM Used memory",
        "VM Memory Swapped",
        "VM Memory Ballooned",
        "#NICs",
        "# vCPUs",
        "vRAM",
        "ESX Version",
        "Vendor",
        "Model",
        "Object ID",
        "UUID",
    ],
    "vDatastore": [
        "Name",
        "Object ID",
        "Type",
        "Hosts",
        "Capacity MiB",
        "Provisioned MiB",
        "In Use MiB",
    ],
    "vSnapshot": [
        "VM",
        "VM UUID",
        "Powerstate",
        "Size MiB (vmsn)",
        "Size MiB (total)",
        "Quiesced",
        "Datacenter",
        "Cluster",
        "Host",
    ],
    "vPartition": ["VM", "VM UUID", "Capacity MiB", "Consumed MiB"],
    "vMemory": ["VM", "VM UUID", "Size MiB", "Reservation"],
    "vDisk": ["VM", "VM UUID", "Shared Bus", "Controller"],
    "vCD": ["VM", "VM UUID", "Powerstate", "Device Type", "Connected"],
    "vUSB": ["VM", "VM UUID", "Powerstate", "Device Type", "Connected"],
    "vNetwork": ["VM", "VM UUID", "Switch", "Connected"],
    "dvPort": [
        "Object ID",
        "Port",
        "Switch",
        "Type",
        "VLAN",
        "Allow Promiscuous",
        "Mac changes",
        "Forged Transmits",
    ],
}

BOOLEAN_HEADERS = {
    "in Maintenance Mode",
    "in Quarantine Mode",
    "Quiesced",
    "Connected",
    "Allow Promiscuous",
    "Mac changes",
    "Forged Transmits",
}

INTEGER_HEADERS = {
    "CPUs",
    "Memory",
    "Provisioned MiB",
    "In use MiB",
    "Speed",
    "#CPU",
    "Cores per CPU",
    "# Cores",
    "# Memory",
    "VM Used memory",
    "VM Memory Swapped",
    "VM Memory Ballooned",
    "#NICs",
    "# vCPUs",
    "vRAM",
    "Capacity MiB",
    "In Use MiB",
    "Size MiB (vmsn)",
    "Size MiB (total)",
    "Consumed MiB",
    "Size MiB",
    "Reservation",
    "Port",
    "VLAN",
}


def normalize_name(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 42)


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def find_sheet(workbook, expected_name: str):
    if expected_name in workbook.sheetnames:
        return workbook[expected_name]

    expected_normalized = normalize_name(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_name(sheet_name) == expected_normalized:
            return workbook[sheet_name]
    return None


def build_header_index(ws) -> dict[str, int]:
    return {
        normalize_name(cell.value): index
        for index, cell in enumerate(ws[1])
        if cell.value not in (None, "")
    }


def coerce_boolean(value: object) -> object:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return "true" if value else "false"

    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1", "connected"}:
        return "true"
    if text in {"false", "no", "n", "0", "disconnected"}:
        return "false"
    return value


def coerce_integer(value: object) -> object:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)

    text = str(value).strip().replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.0+)?", text):
        return int(float(text))
    return value


def coerce_value(header: str, value: object) -> object:
    if header in BOOLEAN_HEADERS:
        return coerce_boolean(value)
    if header in INTEGER_HEADERS:
        return coerce_integer(value)
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def extract_rows(source_sheet, expected_headers: list[str]):
    header_index = build_header_index(source_sheet)
    missing_headers = []
    expected_indices = []
    for header in expected_headers:
        normalized = normalize_name(header)
        if normalized in header_index:
            expected_indices.append(header_index[normalized])
        else:
            expected_indices.append(None)
            missing_headers.append(header)

    rows = []
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        output_row = []
        has_value = False
        for header, source_index in zip(expected_headers, expected_indices):
            value = (
                None
                if source_index is None or source_index >= len(row)
                else row[source_index]
            )
            value = coerce_value(header, value)
            if value not in (None, ""):
                has_value = True
            output_row.append(value)
        if has_value:
            rows.append(output_row)

    return rows, missing_headers


def load_template_headers(template_path: Path) -> dict[str, list[str]]:
    if not template_path.exists():
        return SHEET_SCHEMA

    workbook = load_workbook(template_path, read_only=True, data_only=True)
    try:
        template_headers = {}
        for sheet_name, fallback_headers in SHEET_SCHEMA.items():
            ws = find_sheet(workbook, sheet_name)
            if ws is None:
                template_headers[sheet_name] = fallback_headers
                continue
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            template_headers[sheet_name] = [
                str(value).strip() for value in headers if value not in (None, "")
            ]
        return template_headers
    finally:
        workbook.close()


def rebuild_workbook(source_path: Path, output_path: Path, template_path: Path):
    template_headers = load_template_headers(template_path)
    source_workbook = load_workbook(source_path, read_only=True, data_only=True)

    report = {
        "missing_sheets": [],
        "missing_headers": {},
        "row_counts": {},
    }

    try:
        output_workbook = Workbook()
        output_workbook.remove(output_workbook.active)

        for sheet_name, expected_headers in template_headers.items():
            source_sheet = find_sheet(source_workbook, sheet_name)
            output_sheet = output_workbook.create_sheet(sheet_name)
            output_sheet.append(expected_headers)
            style_header(output_sheet)

            if source_sheet is None:
                report["missing_sheets"].append(sheet_name)
                auto_fit_columns(output_sheet)
                continue

            rows, missing_headers = extract_rows(source_sheet, expected_headers)
            report["missing_headers"][sheet_name] = missing_headers
            report["row_counts"][sheet_name] = len(rows)

            for row in rows:
                output_sheet.append(row)

            auto_fit_columns(output_sheet)

        output_workbook.save(output_path)
    finally:
        source_workbook.close()

    return report


def default_output_path(source_path: Path) -> Path:
    return source_path.with_name(f"{source_path.stem}_azure_migrate_repaired.xlsx")


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Rebuild an RVTools workbook into the strict Azure Migrate import shape. "
            "The output keeps only the required RVTools sheets and rewrites their headers "
            "to the expected schema."
        )
    )
    parser.add_argument(
        "source", type=Path, help="Path to the incoming RVTools workbook"
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Path for the corrected workbook copy",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help="Strict template workbook used to confirm the expected headers",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    source_path = args.source.resolve()
    output_path = (
        args.output.resolve() if args.output else default_output_path(source_path)
    )
    template_path = args.template.resolve()

    if not source_path.exists():
        raise SystemExit(f"Source workbook not found: {source_path}")

    report = rebuild_workbook(source_path, output_path, template_path)

    print(f"Wrote repaired workbook: {output_path}")
    for sheet_name in SHEET_SCHEMA:
        row_count = report["row_counts"].get(sheet_name, 0)
        print(f"- {sheet_name}: {row_count} data rows")

    if report["missing_sheets"]:
        print("Missing source sheets:")
        for sheet_name in report["missing_sheets"]:
            print(f"- {sheet_name}")

    missing_header_sheets = {
        sheet_name: headers
        for sheet_name, headers in report["missing_headers"].items()
        if headers
    }
    if missing_header_sheets:
        print("Missing source headers:")
        for sheet_name, headers in missing_header_sheets.items():
            print(f"- {sheet_name}: {', '.join(headers)}")


if __name__ == "__main__":
    main()
