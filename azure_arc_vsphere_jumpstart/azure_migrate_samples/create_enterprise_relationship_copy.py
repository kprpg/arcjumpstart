from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
ENTERPRISE_SOURCE = BASE_DIR / "azure_migrate_vmware_enterprise_import_sample.xlsx"
ENTERPRISE_TARGET = (
    BASE_DIR
    / "azure_migrate_vmware_enterprise_import_sample_with_app_relationships.xlsx"
)
ENTERPRISE_APP_JSON = BASE_DIR / "enterprise_app_inventory_sample.json"
ENTERPRISE_TECH_CSV = BASE_DIR / "enterprise_technology_inventory_sample.csv"
ENTERPRISE_DEP_CSV = BASE_DIR / "enterprise_dependency_map_sample.csv"
ENTERPRISE_GROUPS_CSV = BASE_DIR / "enterprise_workload_groups_sample.csv"
ENTERPRISE_CHECKLIST_CSV = BASE_DIR / "enterprise_azure_migrate_group_checklist.csv"
ENTERPRISE_MEMBERSHIP_CSV = BASE_DIR / "enterprise_azure_migrate_group_members.csv"
HEALTHCARE_SOURCE = BASE_DIR / "healthcare_payor_provider_vmware_import_sample.xlsx"
HEALTHCARE_TARGET = (
    BASE_DIR
    / "healthcare_payor_provider_vmware_import_sample_with_app_relationships.xlsx"
)
WORKSHOP_TARGET = (
    BASE_DIR / "azure_migrate_vmware_enterprise_sql_sap_citrix_workshop.xlsx"
)
WORKSHOP_CHECKLIST_CSV = BASE_DIR / "azure_migrate_sql_sap_citrix_group_checklist.csv"
WORKSHOP_MEMBERSHIP_CSV = BASE_DIR / "azure_migrate_sql_sap_citrix_group_members.csv"
HEALTHCARE_APP_JSON = BASE_DIR / "healthcare_payor_provider_app_inventory_sample.json"
HEALTHCARE_TECH_CSV = (
    BASE_DIR / "healthcare_payor_provider_technology_inventory_sample.csv"
)
HEALTHCARE_DEP_CSV = BASE_DIR / "healthcare_payor_provider_dependency_map_sample.csv"
HEALTHCARE_GROUPS_CSV = (
    BASE_DIR / "healthcare_payor_provider_workload_groups_sample.csv"
)
HEALTHCARE_CHECKLIST_CSV = BASE_DIR / "healthcare_azure_migrate_group_checklist.csv"
HEALTHCARE_MEMBERSHIP_CSV = BASE_DIR / "healthcare_azure_migrate_group_members.csv"

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
TIER_FILLS = {
    "web": PatternFill("solid", fgColor="E8F1FB"),
    "application": PatternFill("solid", fgColor="EAF7EA"),
    "database": PatternFill("solid", fgColor="FDEBD0"),
    "integration": PatternFill("solid", fgColor="F5E9FB"),
    "euc": PatternFill("solid", fgColor="FCE4EC"),
    "compute": PatternFill("solid", fgColor="E8F8F5"),
    "other": PatternFill("solid", fgColor="F2F4F4"),
}
PATTERN_FILLS = {
    "ha multi-tier application": PatternFill("solid", fgColor="FFF2CC"),
    "sap landscape": PatternFill("solid", fgColor="D5E8D4"),
    "citrix / vdi control plane": PatternFill("solid", fgColor="F8CECC"),
    "three-tier application": PatternFill("solid", fgColor="DAE8FC"),
    "n-tier application": PatternFill("solid", fgColor="E1D5E7"),
}


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        col = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col].width = min(max(max_length + 2, 14), 42)


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def parse_vm_name(vm_name: str) -> tuple[str, str, str, str]:
    parts = vm_name.split("-")
    environment = parts[1]
    instance = parts[-1]
    tier = parts[-2]
    application_name = "-".join(parts[2:-2])
    return environment, application_name, tier, instance


def role_for_tier(tier: str) -> str:
    mapping = {
        "web": "web host / ingress",
        "app": "application tier",
        "api": "api tier",
        "db": "database tier",
        "dbprima": "database primary",
        "dbseco": "database secondary",
        "witnes": "cluster witness",
        "sapweb": "sap web dispatcher",
        "sapapp": "sap application tier",
        "batch": "batch / background processing",
        "contro": "citrix delivery controller",
        "storef": "citrix storefront",
        "worker": "worker / session host",
        "presen": "presentation tier",
        "report": "reporting tier",
        "broker": "integration broker",
        "idm": "identity service",
        "ci": "ci server",
        "artifa": "artifact service",
        "etl": "etl tier",
        "comput": "compute tier",
        "licens": "license service",
    }
    return mapping.get(tier, tier)


def read_csv_rows(path: Path):
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv_rows(path: Path, headers: list[str], rows: list[list]):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def add_sheet(workbook, title: str, headers: list[str], rows: list[list]):
    if title in workbook.sheetnames:
        del workbook[title]
    ws = workbook.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws)
    auto_fit_columns(ws)
    ws.freeze_panes = "A2"
    return ws


def tier_bucket(tier: str) -> str:
    if tier in {"web", "presen", "sapweb", "storef"}:
        return "web"
    if tier in {
        "app",
        "api",
        "sapapp",
        "broker",
        "idm",
        "contro",
        "artifa",
        "ci",
        "licens",
    }:
        return "application"
    if tier in {"db", "dbprima", "dbseco", "witnes"}:
        return "database"
    if tier in {"worker"}:
        return "euc"
    if tier in {"etl", "comput", "batch", "report"}:
        return "compute"
    return "other"


def slugify(value: str) -> str:
    return value.lower().replace(" ", "-").replace("_", "-")


def group_name_for_app(environment: str, criticality: str, app_name: str) -> str:
    criticality_slug = slugify(criticality).replace("-", "")
    return f"grp-{slugify(environment)}-{criticality_slug}-{slugify(app_name)}"


def explicit_cluster_type(app_name: str, tier: str, members: list[str]) -> str:
    if tier in {"dbprima", "dbseco"} or "sqlao" in app_name:
        return "SQL Always On / AG pair"
    if (
        tier == "db"
        and len(members) == 2
        and any(
            "sql" in member or "member-dotnet" in member or "ehr" in member
            for member in members
        )
    ):
        return "SQL cluster / AG pair"
    if tier == "db":
        return "database cluster / HA tier"
    if tier in {"web", "presen"}:
        return "IIS web farm / load-balanced front end"
    if tier in {"app", "api"} and "java" in app_name:
        return "three-tier Java application cluster"
    if tier in {"app", "api"} and "dotnet" in app_name:
        return "three-tier .NET application cluster"
    if tier == "sapapp":
        return "SAP application server cluster"
    if tier == "sapweb":
        return "SAP Web Dispatcher pair"
    if tier == "contro":
        return "Citrix Delivery Controller pair"
    if tier == "storef":
        return "Citrix StoreFront pair"
    if tier == "worker":
        return "Citrix worker pool / session hosts"
    if tier in {"broker", "idm", "report", "batch", "etl", "comput"}:
        return "scaled service tier"
    return "application service cluster"


def inferred_pattern_for_app(
    app_name: str, web_hosts: int, app_hosts: int, db_hosts: int
) -> str:
    lowered = app_name.lower()
    if "sap" in lowered:
        return "SAP landscape"
    if "citrix" in lowered:
        return "Citrix / VDI control plane"
    if web_hosts > 0 and app_hosts > 0 and db_hosts > 0:
        if "java" in lowered or "dotnet" in lowered:
            return "Three-tier application"
        return "HA multi-tier application" if db_hosts >= 2 else "N-tier application"
    if db_hosts >= 2:
        return "HA multi-tier application"
    return "N-tier application"


def style_body_sheet(ws, title: str):
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if title == "ServerApplicationMap" and "tier" in header_map:
            tier_value = ws.cell(row=row[0].row, column=header_map["tier"]).value or ""
            fill = TIER_FILLS[tier_bucket(str(tier_value))]
            for cell in row:
                cell.fill = fill
        elif title == "ApplicationTopology" and "inferred_pattern" in header_map:
            pattern = str(
                ws.cell(row=row[0].row, column=header_map["inferred_pattern"]).value
                or ""
            ).lower()
            fill = PATTERN_FILLS.get(pattern)
            if fill:
                for cell in row:
                    cell.fill = fill
        elif title == "ClusterRelationships" and "cluster_type" in header_map:
            cluster_type = str(
                ws.cell(row=row[0].row, column=header_map["cluster_type"]).value or ""
            ).lower()
            if "sql" in cluster_type or "database" in cluster_type:
                fill = TIER_FILLS["database"]
            elif (
                "iis" in cluster_type
                or "front end" in cluster_type
                or "dispatcher" in cluster_type
            ):
                fill = TIER_FILLS["web"]
            elif "citrix" in cluster_type or "session" in cluster_type:
                fill = TIER_FILLS["euc"]
            elif (
                "sap" in cluster_type
                or "application" in cluster_type
                or "service" in cluster_type
            ):
                fill = TIER_FILLS["application"]
            else:
                fill = TIER_FILLS["other"]
            for cell in row:
                cell.fill = fill


def add_legend_sheet(workbook):
    legend_rows = [
        ["section", "label", "meaning"],
        ["Server/Tier color", "blue", "web / ingress / presentation tiers"],
        ["Server/Tier color", "green", "application / api / middleware tiers"],
        ["Server/Tier color", "orange", "database or database witness tiers"],
        ["Server/Tier color", "pink", "Citrix / EUC worker and session-host tiers"],
        ["Server/Tier color", "teal", "batch, ETL, compute, and reporting tiers"],
        ["Pattern color", "yellow", "HA multi-tier application"],
        ["Pattern color", "light green", "SAP landscape"],
        ["Pattern color", "rose", "Citrix / VDI control plane"],
        ["Pattern color", "light blue", "three-tier application"],
        [
            "Import note",
            "RVTools tabs",
            "Only the original RVTools tabs should be treated as Azure Migrate import candidates.",
        ],
        [
            "Import note",
            "overlay sheets",
            "Application relationship sheets are planning overlays, not a supported Azure Migrate application-dependency import format.",
        ],
    ]
    ws = add_sheet(workbook, "Legend", legend_rows[0], legend_rows[1:])
    fills = {
        2: [
            TIER_FILLS["web"],
            TIER_FILLS["application"],
            TIER_FILLS["database"],
            TIER_FILLS["euc"],
            TIER_FILLS["compute"],
            PATTERN_FILLS["ha multi-tier application"],
            PATTERN_FILLS["sap landscape"],
            PATTERN_FILLS["citrix / vdi control plane"],
            PATTERN_FILLS["three-tier application"],
        ]
    }
    for idx, fill in enumerate(fills[2], start=2):
        for cell in ws[idx]:
            cell.fill = fill
    return ws


def build_workload_groups_rows(app_members, tier_members, app_by_name, group_by_app):
    rows = []
    ordered_tier_names = [
        "web",
        "presen",
        "sapweb",
        "storef",
        "app",
        "api",
        "sapapp",
        "broker",
        "contro",
        "idm",
        "batch",
        "etl",
        "comput",
        "report",
        "dbprima",
        "dbseco",
        "db",
        "witnes",
        "worker",
        "ci",
        "artifa",
        "licens",
    ]
    for app_name, members in sorted(app_members.items()):
        grouped = group_by_app.get(app_name, {})
        app_meta = app_by_name.get(app_name, {})
        tier_entries = []
        for tier in ordered_tier_names:
            tier_members_for_app = tier_members.get((app_name, tier), [])
            if tier_members_for_app:
                tier_entries.append(f"{tier}: {', '.join(tier_members_for_app)}")
        rows.append(
            [
                app_name,
                grouped.get("environment", app_meta.get("environment", "")),
                grouped.get("criticality", app_meta.get("criticality", "")),
                grouped.get("owner", app_meta.get("business_owner", "")),
                grouped.get("vm_count", len(members)),
                app_meta.get("target_pattern", grouped.get("target_pattern", "")),
                app_meta.get("notes", ""),
                " | ".join(tier_entries),
                "; ".join(members),
            ]
        )
    return rows


def build_portal_checklist_rows(
    app_members,
    tier_members,
    app_by_name,
    group_by_app,
    tech_by_vm,
):
    rows = []
    for app_name, members in sorted(app_members.items()):
        grouped = group_by_app.get(app_name, {})
        app_meta = app_by_name.get(app_name, {})
        environment = grouped.get("environment", app_meta.get("environment", ""))
        criticality = grouped.get("criticality", app_meta.get("criticality", ""))
        owner = grouped.get("owner", app_meta.get("business_owner", ""))
        group_name = group_name_for_app(environment, criticality, app_name)
        tiers = sorted(tier for name, tier in tier_members.keys() if name == app_name)
        waves = sorted(
            {
                tech_by_vm.get(member, {}).get("migration_wave", "")
                for member in members
                if tech_by_vm.get(member, {}).get("migration_wave", "")
            }
        )
        target_hints = sorted(
            {
                tech_by_vm.get(member, {}).get("suggested_target", "")
                for member in members
                if tech_by_vm.get(member, {}).get("suggested_target", "")
            }
        )
        rows.append(
            [
                group_name,
                app_name,
                environment,
                criticality,
                owner,
                len(members),
                "; ".join(tiers),
                "; ".join(waves),
                "; ".join(target_hints)
                or app_meta.get("target_pattern", grouped.get("target_pattern", "")),
                "; ".join(members),
                "Discovery and assessment > Servers, databases and web apps > Groups",
                "Create one assessment group per application, verify tier completeness, then run sizing and readiness assessments.",
            ]
        )
    return rows


def build_group_membership_rows(server_details, app_by_name, group_by_app):
    rows = []
    for detail in sorted(
        server_details,
        key=lambda item: (
            item["environment"],
            item["application_name"],
            item["tier"],
            item["vm_name"],
        ),
    ):
        app_name = detail["application_name"]
        app_meta = app_by_name.get(app_name, {})
        criticality = detail["criticality"] or app_meta.get("criticality", "")
        rows.append(
            [
                group_name_for_app(detail["environment"], criticality, app_name),
                app_name,
                detail["vm_name"],
                detail["tier"],
                detail["role"],
                detail["environment"],
                criticality,
                detail["suggested_target"],
                detail["migration_wave"],
                detail["technologies"],
                app_meta.get(
                    "technical_owner", group_by_app.get(app_name, {}).get("owner", "")
                ),
            ]
        )
    return rows


def workshop_focus_rows(topology_rows, cluster_rows, dependency_rows):
    def include_app(name: str) -> bool:
        lowered = name.lower()
        return any(
            token in lowered
            for token in ["sap", "citrix", "sqlao", "ehr", "member-dotnet", "payments"]
        )

    focus_summary = []
    for row in topology_rows:
        if include_app(row[0]):
            focus_summary.append([row[0], row[8], row[9], row[10]])

    focus_clusters = [row for row in cluster_rows if include_app(row[0])]
    focus_deps = [
        row
        for row in dependency_rows
        if include_app(row[0].split(":", 1)[0]) or include_app(row[1].split(":", 1)[0])
    ]
    return focus_summary, focus_clusters, focus_deps


def build_relationship_copy(
    source: Path,
    target: Path,
    app_json: Path,
    tech_csv: Path,
    dep_csv: Path,
    groups_csv: Path,
    checklist_csv: Path,
    membership_csv: Path,
    include_workshop_focus: bool = False,
):
    wb = load_workbook(source)
    vinfo = wb["vInfo"]

    app_inventory_payload = json.loads(app_json.read_text(encoding="utf-8"))
    app_inventory = app_inventory_payload["applications"]
    app_by_name = {item["application_name"]: item for item in app_inventory}
    tech_rows = read_csv_rows(tech_csv)
    tech_by_vm = {row["vm"]: row for row in tech_rows}
    dep_rows = read_csv_rows(dep_csv)
    group_rows = read_csv_rows(groups_csv)
    group_by_app = {row["application_name"]: row for row in group_rows}

    server_rows = []
    server_details = []
    app_members = defaultdict(list)
    tier_members = defaultdict(list)

    for row in vinfo.iter_rows(min_row=2, values_only=True):
        vm_name = row[0]
        os_name = row[7]
        environment, app_name, tier, instance = parse_vm_name(vm_name)
        tech = tech_by_vm.get(vm_name, {})
        app_members[app_name].append(vm_name)
        tier_members[(app_name, tier)].append(vm_name)
        criticality = app_by_name.get(app_name, {}).get(
            "criticality", group_by_app.get(app_name, {}).get("criticality", "")
        )
        role = role_for_tier(tier)
        server_rows.append(
            [
                vm_name,
                environment,
                app_name,
                tier,
                role,
                instance,
                os_name,
                tech.get("technologies", ""),
                tech.get("suggested_target", ""),
                tech.get("migration_wave", ""),
            ]
        )
        server_details.append(
            {
                "vm_name": vm_name,
                "environment": environment,
                "application_name": app_name,
                "tier": tier,
                "role": role,
                "criticality": criticality,
                "instance": instance,
                "os": os_name,
                "technologies": tech.get("technologies", ""),
                "suggested_target": tech.get("suggested_target", ""),
                "migration_wave": tech.get("migration_wave", ""),
            }
        )

    topology_rows = []
    for app_name, members in sorted(app_members.items()):
        grouped = group_by_app.get(app_name, {})
        app_meta = app_by_name.get(app_name, {})
        web_hosts = sum(
            len(tier_members.get((app_name, tier), []))
            for tier in {"web", "presen", "sapweb", "storef"}
        )
        app_hosts = sum(
            len(tier_members.get((app_name, tier), []))
            for tier in {
                "app",
                "api",
                "sapapp",
                "broker",
                "contro",
                "storef",
                "idm",
                "batch",
                "etl",
                "comput",
                "report",
            }
        )
        db_hosts = sum(
            len(tier_members.get((app_name, tier), []))
            for tier in {"db", "dbprima", "dbseco", "witnes"}
        )
        inferred_pattern = inferred_pattern_for_app(
            app_name, web_hosts, app_hosts, db_hosts
        )

        ha_notes = []
        if db_hosts >= 2:
            if "sqlao" in app_name or any(
                key[0] == app_name and key[1] in {"dbprima", "dbseco"}
                for key in tier_members
            ):
                ha_notes.append(
                    "Database members can be represented explicitly as a SQL Always On / AG pair."
                )
            else:
                ha_notes.append(
                    "Database tier has multiple nodes and can be modeled as a SQL cluster / AG pair or HA database set."
                )
        if web_hosts >= 2:
            ha_notes.append(
                "Front-end tier can be documented as an IIS web farm, presentation farm, or SAP/Citrix ingress pair."
            )
        if app_hosts >= 2:
            ha_notes.append(
                "Application tier supports active-active or scaled-out middle-tier grouping."
            )

        topology_rows.append(
            [
                app_name,
                grouped.get("environment", app_meta.get("environment", "")),
                grouped.get("criticality", app_meta.get("criticality", "")),
                grouped.get("owner", app_meta.get("business_owner", "")),
                len(members),
                web_hosts,
                app_hosts,
                db_hosts,
                inferred_pattern,
                app_meta.get("target_pattern", grouped.get("target_pattern", "")),
                " ".join(ha_notes) or "Single-node tiers where present.",
            ]
        )

    cluster_rows = []
    for (app_name, tier), members in sorted(tier_members.items()):
        if len(members) < 2:
            continue
        cluster_rows.append(
            [
                app_name,
                tier,
                explicit_cluster_type(app_name, tier, members),
                len(members),
                "; ".join(members),
                "Inferred from repeated VM naming on the same logical tier and enriched with explicit interpretation.",
            ]
        )

    dependency_rows = []
    for row in dep_rows:
        source_app, source_tier = row["source"].split(":", 1)
        target_app, target_tier = row["target"].split(":", 1)
        dependency_rows.append(
            [
                row["source"],
                row["target"],
                row["protocol"],
                row["port"],
                row["interaction_type"],
                row["criticality"],
                "; ".join(tier_members.get((source_app, source_tier), [])),
                "; ".join(tier_members.get((target_app, target_tier), [])),
            ]
        )

    overlay_notes = [
        [
            "Purpose",
            "This copy preserves the RVTools tabs exactly and adds planning-only sheets for application grouping, N-tier topology, explicit cluster labels, and dependencies.",
        ],
        [
            "Import caution",
            "Azure Migrate file import supports RVTools-style inventory only. These overlay sheets are for planning and documentation, not for a supported application-dependency import path.",
        ],
        [
            "What changed",
            "No required RVTools sheet names were changed. Added sheets: OverlayNotes, Legend, ServerApplicationMap, ApplicationTopology, ClusterRelationships, ApplicationDependencies, WorkloadGroupsForAzureMigrate, AzureMigrateGroupChecklist.",
        ],
        [
            "Example interpretation",
            "Two DB servers can be documented as SQL Cluster or SQL Always On / AG members. Multiple IIS or presentation servers can be documented as a web farm. Citrix controller, StoreFront, and worker pools can be grouped explicitly.",
        ],
    ]

    workload_group_rows = build_workload_groups_rows(
        app_members, tier_members, app_by_name, group_by_app
    )
    checklist_rows = build_portal_checklist_rows(
        app_members, tier_members, app_by_name, group_by_app, tech_by_vm
    )
    group_membership_rows = build_group_membership_rows(
        server_details, app_by_name, group_by_app
    )

    add_sheet(wb, "OverlayNotes", ["key", "value"], overlay_notes)
    add_legend_sheet(wb)
    server_ws = add_sheet(
        wb,
        "ServerApplicationMap",
        [
            "vm_name",
            "environment",
            "application_name",
            "tier",
            "role",
            "instance",
            "os",
            "technologies",
            "suggested_target",
            "migration_wave",
        ],
        server_rows,
    )
    topology_ws = add_sheet(
        wb,
        "ApplicationTopology",
        [
            "application_name",
            "environment",
            "criticality",
            "owner",
            "server_count",
            "web_hosts",
            "app_hosts",
            "db_hosts",
            "inferred_pattern",
            "target_pattern",
            "ha_notes",
        ],
        topology_rows,
    )
    cluster_ws = add_sheet(
        wb,
        "ClusterRelationships",
        [
            "application_name",
            "tier",
            "cluster_type",
            "member_count",
            "members",
            "notes",
        ],
        cluster_rows,
    )
    dependency_ws = add_sheet(
        wb,
        "ApplicationDependencies",
        [
            "source_group",
            "target_group",
            "protocol",
            "port",
            "interaction_type",
            "criticality",
            "source_members",
            "target_members",
        ],
        dependency_rows,
    )
    workload_groups_ws = add_sheet(
        wb,
        "WorkloadGroupsForAzureMigrate",
        [
            "application_name",
            "environment",
            "criticality",
            "owner",
            "vm_count",
            "target_pattern",
            "notes",
            "tier_membership",
            "all_members",
        ],
        workload_group_rows,
    )
    checklist_ws = add_sheet(
        wb,
        "AzureMigrateGroupChecklist",
        [
            "group_name",
            "application_name",
            "environment",
            "criticality",
            "owner",
            "vm_count",
            "included_tiers",
            "migration_waves",
            "target_hints",
            "all_members",
            "portal_path",
            "checklist_note",
        ],
        checklist_rows,
    )

    if include_workshop_focus:
        focus_summary, focus_clusters, focus_deps = workshop_focus_rows(
            topology_rows, cluster_rows, dependency_rows
        )
        add_sheet(
            wb,
            "WorkshopNotes",
            ["key", "value"],
            [
                [
                    "scope",
                    "Focused workshop view for SQL clusters, SQL Always On, SAP, Citrix, and similar high-attention migration patterns.",
                ],
                [
                    "how_to_use",
                    "Use this workbook during technical workshops to walk the team through cluster semantics, dependency chains, and migration-wave sequencing for the hardest patterns first.",
                ],
            ],
        )
        workshop_summary_ws = add_sheet(
            wb,
            "WorkshopFocusSummary",
            [
                "application_name",
                "inferred_pattern",
                "target_pattern",
                "ha_notes",
            ],
            focus_summary,
        )
        workshop_clusters_ws = add_sheet(
            wb,
            "WorkshopFocusClusters",
            [
                "application_name",
                "tier",
                "cluster_type",
                "member_count",
                "members",
                "notes",
            ],
            focus_clusters,
        )
        workshop_deps_ws = add_sheet(
            wb,
            "WorkshopFocusDependencies",
            [
                "source_group",
                "target_group",
                "protocol",
                "port",
                "interaction_type",
                "criticality",
                "source_members",
                "target_members",
            ],
            focus_deps,
        )

    for title, ws in {
        "ServerApplicationMap": server_ws,
        "ApplicationTopology": topology_ws,
        "ClusterRelationships": cluster_ws,
        "ApplicationDependencies": dependency_ws,
        "WorkloadGroupsForAzureMigrate": workload_groups_ws,
        "AzureMigrateGroupChecklist": checklist_ws,
    }.items():
        style_body_sheet(ws, title)

    if include_workshop_focus:
        for title, ws in {
            "WorkshopFocusSummary": workshop_summary_ws,
            "WorkshopFocusClusters": workshop_clusters_ws,
            "WorkshopFocusDependencies": workshop_deps_ws,
        }.items():
            style_body_sheet(
                ws,
                "ApplicationTopology"
                if title == "WorkshopFocusSummary"
                else "ClusterRelationships"
                if title == "WorkshopFocusClusters"
                else "ApplicationDependencies",
            )

    wb.save(target)
    write_csv_rows(
        checklist_csv,
        [
            "group_name",
            "application_name",
            "environment",
            "criticality",
            "owner",
            "vm_count",
            "included_tiers",
            "migration_waves",
            "target_hints",
            "all_members",
            "portal_path",
            "checklist_note",
        ],
        checklist_rows,
    )
    write_csv_rows(
        membership_csv,
        [
            "group_name",
            "application_name",
            "vm_name",
            "tier",
            "role",
            "environment",
            "criticality",
            "suggested_target",
            "migration_wave",
            "technologies",
            "technical_owner",
        ],
        group_membership_rows,
    )
    print(f"Created {target}")
    print(f"Created {checklist_csv}")
    print(f"Created {membership_csv}")


def main():
    build_relationship_copy(
        ENTERPRISE_SOURCE,
        ENTERPRISE_TARGET,
        ENTERPRISE_APP_JSON,
        ENTERPRISE_TECH_CSV,
        ENTERPRISE_DEP_CSV,
        ENTERPRISE_GROUPS_CSV,
        ENTERPRISE_CHECKLIST_CSV,
        ENTERPRISE_MEMBERSHIP_CSV,
    )
    build_relationship_copy(
        ENTERPRISE_SOURCE,
        WORKSHOP_TARGET,
        ENTERPRISE_APP_JSON,
        ENTERPRISE_TECH_CSV,
        ENTERPRISE_DEP_CSV,
        ENTERPRISE_GROUPS_CSV,
        WORKSHOP_CHECKLIST_CSV,
        WORKSHOP_MEMBERSHIP_CSV,
        include_workshop_focus=True,
    )
    build_relationship_copy(
        HEALTHCARE_SOURCE,
        HEALTHCARE_TARGET,
        HEALTHCARE_APP_JSON,
        HEALTHCARE_TECH_CSV,
        HEALTHCARE_DEP_CSV,
        HEALTHCARE_GROUPS_CSV,
        HEALTHCARE_CHECKLIST_CSV,
        HEALTHCARE_MEMBERSHIP_CSV,
    )


if __name__ == "__main__":
    main()
